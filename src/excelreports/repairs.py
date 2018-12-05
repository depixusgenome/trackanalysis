#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"read-write existing files"
from    typing          import Dict, List
import  os
import  re
import  shutil
import  tempfile

from    zipfile         import ZipFile, ZIP_STORED, ZipInfo
from    xml.etree       import ElementTree
from    pandas          import DataFrame as _DataFrame

import  numpy
import  openpyxl
from    openpyxl.utils  import column_index_from_string as getcolindex

class UpdateableZipFile(ZipFile):
    """
    Add delete (via remove_file) and update (via writestr and write methods) To
    enable update features use UpdateableZipFile with the 'with statement',
    Upon  __exit__ (if updates were applied) a new zip file will override the
    exiting one with the updates
    """

    DeleteMarker = type('DeleteMarker', tuple(), dict())

    def __init__(self, file, mode="r", compression=ZIP_STORED, allowZip64=False):
        # Init base
        super(UpdateableZipFile, self).__init__(file, mode=mode,
                                                compression=compression,
                                                allowZip64=allowZip64)
        # track file to override in zip
        self._replace: dict = {}
        # Whether the with statement was called
        self._allow_updates = False

    def writestr(self, zinfo_or_arcname, # pylint: disable=arguments-differ
                 byts, compress_type=None):
        if isinstance(zinfo_or_arcname, ZipInfo):
            name = zinfo_or_arcname.filename
        else:
            name = zinfo_or_arcname
        # If the file exits, and needs to be overridden,
        # mark the entry, and create a temp-file for it
        # we allow this only if the with statement is used
        if self._allow_updates and name in self.namelist():
            temp_file = self._replace[name] = self._replace.get(name,
                                                                tempfile.TemporaryFile())
            temp_file.write(byts)
        # Otherwise just act normally
        else:
            super(UpdateableZipFile, self).writestr(zinfo_or_arcname,
                                                    byts, compress_type=compress_type)

    def write(self, filename, arcname=None, compress_type=None): # pylint: disable=arguments-differ
        arcname = arcname or filename
        # If the file exits, and needs to be overridden,
        # mark the entry, and create a temp-file for it
        # we allow this only if the with statement is used
        if self._allow_updates and arcname in self.namelist():
            temp_file = self._replace[arcname] = self._replace.get(arcname,
                                                                   tempfile.TemporaryFile())
            with open(filename, "rb") as source:
                shutil.copyfileobj(source, temp_file)
        # Otherwise just act normally
        else:
            super(UpdateableZipFile, self).write(filename,
                                                 arcname=arcname, compress_type=compress_type)

    def __enter__(self):
        # Allow updates
        self._allow_updates = True
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        # call base to close zip file, organically
        try:
            super(UpdateableZipFile, self).__exit__(exc_type, exc_val, exc_tb)
            if len(self._replace) > 0:
                self._rebuild_zip()
        finally:
            # In case rebuild zip failed,
            # be sure to still release all the temp files
            self._close_all_temp_files()
            self._allow_updates = False

    def _close_all_temp_files(self):
        for temp_file in self._replace.values():
            if hasattr(temp_file, 'close'):
                temp_file.close()

    def remove_file(self, path):
        "removes a file from the archive"
        self._replace[path] = self.DeleteMarker()

    def _rebuild_zip(self):
        tempdir = tempfile.mkdtemp()
        try:
            temp_zip_path = os.path.join(tempdir, 'new.zip')
            with ZipFile(self.filename, 'r') as zip_read:
                # Create new zip with assigned properties
                with ZipFile(temp_zip_path, 'w', compression=self.compression,
                             allowZip64=self._allowZip64) as zip_write:
                    for item in zip_read.infolist():
                        # Check if the file should be replaced / or deleted
                        replacement = self._replace.get(item.filename, None)
                        # If marked for deletion, do not copy file to new zipfile
                        if isinstance(replacement, self.DeleteMarker):
                            del self._replace[item.filename]
                            continue
                        # If marked for replacement, copy temp_file, instead of old file
                        elif replacement is not None:
                            del self._replace[item.filename]
                            # Write replacement to archive,
                            # and then close it (deleting the temp file)
                            replacement.seek(0)
                            data = replacement.read()
                            replacement.close()
                        else:
                            data = zip_read.read(item.filename)
                        zip_write.writestr(item, data)
            # Override the archive with the updated one
            shutil.move(temp_zip_path, self.filename)
        finally:
            shutil.rmtree(tempdir)

class DataFrame(_DataFrame):
    "read-write existing files"
    def __init__(self, fname, sheet, args):
        book = openpyxl.load_workbook(fname, read_only = True)#, data_only = True)
        rows = enumerate(book[sheet].iter_rows())
        for _, row in rows:
            cols = tuple((col.value, i) for i, col in enumerate(row) if col.value in args)
            if len(cols):
                break

        if len(cols) != len(args):
            raise KeyError('missing columns %s in file'
                           % (set(name for name, _ in cols) - set(args)),
                           "warning")

        vals: Dict[str, List] = dict((name, []) for name , iC in cols)
        inds = []
        for i, row in rows:
            if all(col.value is None for col in row):
                break

            for name, j in cols:
                vals[name].append(args[name](row[j].value))
            inds.append(i+1)

        super().__init__(data  = vals, index = inds)
        self._fname = fname
        self._sheet = sheet
        self._args  = args
        self._cols  = cols

    def paste(self, filename, key):
        "pastes values into a file"
        if filename == self._fname:
            return None

        other      = DataFrame(filename, self._sheet, self._args)
        cols       = list(name      for name, _ in  self._cols)
        dcols      = list(name      for name    in cols if name != key)
        icols      = dict((name, _) for name, _ in self._cols)
        diffs: set = set()
        equ        = lambda i, j: i == j or all((numpy.isnan(i), numpy.isnan(j)))
        for elem in frozenset(self[key].unique()) & frozenset(other[key].unique()):
            vother = other.loc[other[key] == elem, dcols]
            vself  = self .loc[self [key] == elem, dcols].values[0]
            diffs.update((vother.index[0], icols[name])
                         for _, name in enumerate(dcols)
                         if not equ(vother.values[0][_], vself[_]))

            other.loc[other[key] == elem, dcols] = self .loc[self [key] == elem, cols]

        return other.replaceinfile(*dcols, diffs = frozenset(diffs))

    def getcolindex(self, name):
        "returns the column index"
        for col, icol in self._cols:
            if col == name:
                return icol
        return None

    def replaceinfile(self, *cols, diffs = None):
        "replaces values in a file"
        if diffs is not None and len(diffs) == 0:
            return
        if isinstance(diffs, dict):
            diffs = frozenset(diffs.items())

        args  = dict(self._cols)
        icols = tuple(args[col] for col in cols)
        cols  = dict((j,i) for i,j in self._cols) # type: ignore

        def _getid():
            book = openpyxl.load_workbook(self._fname,
                                          read_only = True,
                                          data_only = True)
            return book.get_sheet_names().index(self._sheet)+1
        sheet = "xl/worksheets/sheet%d.xml" % _getid()


        def _itercells(tree):
            recol = re.compile(r"([A-Z]+)(\d+)")
            data  = next((item for item in tree if item.tag.endswith("sheetData")), ())

            for col in iter(col for row in data for col in row):
                match = recol.match(col.get("r"))
                if match is None:
                    continue
                irow  = int(match.group(2))

                icol  = getcolindex(match.group(1))-1
                if diffs is not None:
                    if (irow, icol) not in diffs:
                        continue
                elif irow not in self.index or icol not in icols:
                    continue

                yield (icol, irow, col)

        with UpdateableZipFile(self._fname) as stream:
            tree = ElementTree.fromstring(stream.read(sheet))

            for icol, irow, col in  _itercells(tree):
                val = self[cols[icol]][irow]
                if numpy.isnan(val):
                    if len(col):
                        col.remove(col[0])
                else:
                    if len(col) == 0:
                        ElementTree.SubElement(col, "s:v")

                    col[0].text = str(val)

            stream.writestr(sheet, ElementTree.tostring(tree))
