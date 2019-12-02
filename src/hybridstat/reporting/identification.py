#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extracts information from a report
"""
from typing                 import Optional, Sequence, Tuple, List, Dict, Callable
from pathlib                import Path
import warnings

with warnings.catch_warnings():
    warnings.filterwarnings('ignore', category = DeprecationWarning)
    from openpyxl               import load_workbook

from excelreports.creation  import writecolumns

IdType  = Tuple[int,str,Optional[float],Optional[float]]
IdsType = List[IdType]

def _id(row, ibead):
    if len(row) <= ibead:
        return None
    val = row[ibead].value
    if val is None:
        return None

    if isinstance(val, str):
        try:
            return int(val.split('B')[-1]) if 'B' in val else int(val)
        except ValueError:
            return None

    try:
        return int(val)
    except ValueError:
        return None

def _tofloat(row, ibead):
    if len(row) <= ibead:
        return None
    val = row[ibead].value
    if val is not None:
        try:
            return float(val)
        except ValueError:
            pass
    return None

def _add(info, row, ibead, ref):
    if len(row) <= ibead:
        return None
    val = row[ibead].value
    if val is None:
        return None

    if isinstance(val, str):
        try:
            beadid = int(val.split('B')[-1]) if 'B' in val else int(val)
        except ValueError:
            return None

    try:
        beadid = int(val)
    except ValueError:
        return None

    info.setdefault(ref, []).append(beadid)
    return None

def _read_summary(rows) -> IdsType:
    ids: List[Optional[int]] = [None, None, None, None]
    names                    = ('bead', 'reference', 'stretch', 'bias')
    for row in rows:
        for i, cell in enumerate(row):
            cur = str(cell.value).split('(')[0].strip().lower()
            try:
                ids[names.index(cur)] = i
            except ValueError:
                pass
        if ids.count(None) != 4:
            break

    info: IdsType = list()
    if ids.count(None) == 0:
        cnv: Sequence[Callable] = (
            _id,
            lambda r, i: (None if len(r) <= i else str(r[i].value)),
            _tofloat,
            _tofloat
        )
        for row in rows:
            vals = tuple(fcn(row, idx) for fcn, idx in zip(cnv, ids))
            if None not in vals[:2]:
                info.append(vals) # type: ignore
    return info

def _read_identifications(rows) -> IdsType:
    info: Dict[str,List[int]]  = dict()
    inds: List[Tuple[int,str]] = []
    for row in rows:
        for i, cell in enumerate(row):
            val = str(cell.value)
            if val != "":
                inds.append((i, val))
        if len(inds):
            break

    for row in rows:
        for ibead, ref in inds:
            _add(info, row, ibead, ref)

    res = [] # type: IdsType
    for hpin, beads in info.items():
        res.extend((i, hpin, None, None) for i in beads)
    return res

def readparams(fname:str) -> IdsType:
    "extracts bead ids and their reference from a report"
    if not Path(fname).exists():
        raise ValueError("Id file path unreachable","warning")
    if Path(fname).is_dir():
        raise ValueError("Id file path is a directory", "warning")
    wbook = load_workbook(filename=fname, read_only=True)
    for sheetname in wbook.sheetnames:
        if sheetname.lower() == "summary":
            return _read_summary(iter(wbook[sheetname].rows))

    for sheetname in wbook.sheetnames:
        if sheetname.lower() == "identification":
            return _read_identifications(iter(wbook[sheetname].rows))

    for sheetname in wbook.sheetnames:
        return _read_identifications(iter(wbook[sheetname].rows))
    res: IdsType = []
    return res

def writeparams(fname:str, items: Sequence[Tuple[str,Sequence[int]]]):
    "write bead ids and their reference to a report"
    writecolumns(fname, "Identification", items)
